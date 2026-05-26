"""
Build Japan university target matrix Excel report for Celebal SDE intern program.
Output: data/reports/japan-university-target-matrix-2026Q2.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path


OUTPUT = Path(__file__).resolve().parent.parent / "data" / "reports" / "japan-university-target-matrix-2026Q2.xlsx"


COLUMNS = [
    ("University", 28),
    ("City", 10),
    ("Website", 38),
    ("LinkedIn URL", 50),
    ("Targeted Course / Collaborations", 50),
    ("No. of Recruiters\n(Competitive Intensity)", 18),
    ("Feedback (Our Assessment)", 55),
    ("Contact Person", 35),
    ("Email", 35),
    ("Phone", 18),
    ("Possibility of Collaboration", 18),
    ("How (Collaboration Mechanism)", 55),
    ("Past Collaboration Evidence / Link", 50),
    ("Fit Score (1-10)", 10),
    ("Recommended Cycle", 16),
    ("Best First-Contact Channel", 38),
    ("Language Preference", 18),
    ("Off-cycle from Shinsotsu", 14),
    ("English-Medium Engineering", 22),
    ("Stipend Benchmark (JPY/month)", 18),
    ("Risk Notes", 50),
]

ROWS = [
    {
        "University": "Nagoya University\n(名古屋大学 / Meidai)",
        "City": "Nagoya",
        "Website": "https://en.nagoya-u.ac.jp/",
        "LinkedIn URL": "https://www.linkedin.com/school/nagoya-university/",
        "Targeted Course / Collaborations": "G30 International Programs (English-medium UNDERGRAD engineering — unique in Japan); Graduate School of Informatics (CS, Mathematical Informatics, Intelligent Systems)",
        "No. of Recruiters\n(Competitive Intensity)": "Medium",
        "Feedback (Our Assessment)": "⭐ STRATEGIC ANCHOR. Only university in shortlist with named human POC. G30 = pre-validated English fluency. October admission cycle bypasses shinsotsu collision. Institution actively pursuing international MoUs (recent Edinburgh global campus).",
        "Contact Person": "Michelle Kuhn, International Student Advisor, Graduate School of Informatics",
        "Email": "kuhnm@i.nagoya-u.ac.jp",
        "Phone": "+81-52-788-6269",
        "Possibility of Collaboration": "High",
        "How (Collaboration Mechanism)": "5-year MoU via Industry-Academia-Government Collaboration HQ (sangaku@t.mail.nagoya-u.ac.jp). 4 work streams: (1) talent pipeline (interns→FTE), (2) sponsored research / capstones, (3) faculty engagement (guest lectures, workshops), (4) bilateral exchange (Jaipur↔Nagoya residencies).",
        "Past Collaboration Evidence / Link": "Global campus partnership with University of Edinburgh (announced); India-focused engagement events on record; established Toyota/Denso/Aisin industrial partnership tradition. URL: http://www.aip.nagoya-u.ac.jp/en/introduction/index.html",
        "Fit Score (1-10)": 9,
        "Recommended Cycle": "Year 1 (Anchor)",
        "Best First-Contact Channel": "Email Michelle Kuhn directly, cc Career Support Center",
        "Language Preference": "English supported",
        "Off-cycle from Shinsotsu": "Yes (G30 Oct intake)",
        "English-Medium Engineering": "Yes — UNDERGRAD G30",
        "Stipend Benchmark (JPY/month)": "¥150,000–220,000",
        "Risk Notes": "G30 cohort is smaller in pure CS than Automotive Engineering; differentiate vs Toyota-ecosystem gravitational pull.",
    },
    {
        "University": "Institute of Science Tokyo\n(科学大 / formerly Tokyo Tech)",
        "City": "Tokyo",
        "Website": "https://www.isct.ac.jp/en/",
        "LinkedIn URL": "https://www.linkedin.com/school/tokyo-institute-of-technology/ [transitioning post-2024 merger]",
        "Targeted Course / Collaborations": "IGES (International Graduate Program in Engineering & Sciences) Master's/PhD — fully English-medium; School of Computing (CS Dept); Mathematical & Computing Science",
        "No. of Recruiters\n(Competitive Intensity)": "High",
        "Feedback (Our Assessment)": "Highest CS/engineering concentration of any Japanese university. IGES grad cohort is English-fluent, international, and accustomed to global firms. Lab-industry collaboration culture lowers friction for structured long-term internships. Career office has explicit recruiter intake process.",
        "Contact Person": "Not publicly listed (use Career Support Services general contact)",
        "Email": "career.info@ssc.isct.ac.jp",
        "Phone": "+81-3-5734-3012",
        "Possibility of Collaboration": "High",
        "How (Collaboration Mechanism)": "Recruiter registration via Career Support Services + career-tasu employer system (Japanese-only). Optional supplement: institutional partnership via Center for Innovation Management (info@cim.isct.ac.jp).",
        "Past Collaboration Evidence / Link": "Tokyo Tech alumni heavily represented at Google Japan, Amazon Japan, Microsoft Japan, Indeed, Mercari (Japanese tech press cites). Historical collaborations with Toyota, Hitachi, Sony documented. URL: https://www.titech.ac.jp/english/student-support/industry-researchers/student-employment/recruiting",
        "Fit Score (1-10)": 9,
        "Recommended Cycle": "Year 2",
        "Best First-Contact Channel": "Email career.info@ssc.isct.ac.jp; expect JP-language follow-up",
        "Language Preference": "JP primary; EN email OK",
        "Off-cycle from Shinsotsu": "Partial (IGES flexible)",
        "English-Medium Engineering": "Yes — GRADUATE IGES",
        "Stipend Benchmark (JPY/month)": "¥200,000–300,000",
        "Risk Notes": "Post-2024 merger admin churn (URL/POC may change); high gaishi-kei competition; legacy titech.ac.jp domain still active.",
    },
    {
        "University": "Waseda University\n(早稲田大学)",
        "City": "Tokyo",
        "Website": "https://www.waseda.jp/top/en/",
        "LinkedIn URL": "https://www.linkedin.com/school/waseda-university/",
        "Targeted Course / Collaborations": "Faculty of Science and Engineering (FSE/CSE/ASE) English-medium degree programs — Computer Science & Communications Engineering, Computer Science & Engineering English tracks; School of International Liberal Studies (SILS) for adjacent talent.",
        "No. of Recruiters\n(Competitive Intensity)": "Very High",
        "Feedback (Our Assessment)": "Most internationally oriented top-tier Tokyo private. Explicit international-student career support track. Career Center is centralized, contactable, and experienced engaging foreign employers. Largest international student body in Japan.",
        "Contact Person": "Not publicly listed (Career Center general)",
        "Email": "career@list.waseda.jp; intl-fse@list.waseda.jp (FSE direct)",
        "Phone": "+81-3-3203-4332",
        "Possibility of Collaboration": "High",
        "How (Collaboration Mechanism)": "Submit Career Center Visitors / Meeting Request form (https://my.waseda.jp/application/noauth/...) + email intl-fse@list.waseda.jp to engage FSE international-student population directly. No distinct international career sub-track is published; central Career Center is the only path.",
        "Past Collaboration Evidence / Link": "Public alumni record at IBM Japan, Accenture, Deloitte, Goldman Sachs, Amazon Japan, TCS Japan, Infosys Japan. URL: https://www.waseda.jp/inst/career/en/company/",
        "Fit Score (1-10)": 9,
        "Recommended Cycle": "Year 3",
        "Best First-Contact Channel": "Career Center Visitors form + intl-fse@list.waseda.jp",
        "Language Preference": "Both supported",
        "Off-cycle from Shinsotsu": "No",
        "English-Medium Engineering": "Partial (FSE EN tracks)",
        "Stipend Benchmark (JPY/month)": "¥200,000–300,000",
        "Risk Notes": "Highly competitive — students courted heavily by gaishi-kei FAANG-tier; differentiation required. 6-month commitment may collide with shūkatsu prep.",
    },
    {
        "University": "The University of Tokyo\n(東京大学 / Todai)",
        "City": "Tokyo",
        "Website": "https://www.u-tokyo.ac.jp/en/",
        "LinkedIn URL": "https://www.linkedin.com/school/the-university-of-tokyo/",
        "Targeted Course / Collaborations": "Graduate School of Information Science and Technology (English-medium GSC tracks); Graduate School of Engineering. Undergrad engineering is mostly Japanese.",
        "No. of Recruiters\n(Competitive Intensity)": "Very High",
        "Feedback (Our Assessment)": "Top brand in Japan but career office is decentralized (per-faculty); students are heavily over-courted by gaishi-kei FAANG firms, top consulting, and investment banking. SDE intern outreach has highest friction in the shortlist.",
        "Contact Person": "Not publicly listed",
        "Email": "Via central Career Support Office; not centrally published as employer-facing alias",
        "Phone": "+81-3-3814-3911 (main)",
        "Possibility of Collaboration": "Medium",
        "How (Collaboration Mechanism)": "Best path is lab-mediated through individual Faculty of IST PIs — no central employer-friendly channel for foreign tech intern partnerships. Career Support Office available but selective.",
        "Past Collaboration Evidence / Link": "Public placements at Google Japan, IBM Japan, Goldman Sachs, McKinsey, Accenture Japan. URL: https://www.u-tokyo.ac.jp/en/prospective-students/career_support.html",
        "Fit Score (1-10)": 7,
        "Recommended Cycle": "Year 3+ (low priority)",
        "Best First-Contact Channel": "Lab-mediated via GSC English-program faculty PI",
        "Language Preference": "JP primary; EN at GSC",
        "Off-cycle from Shinsotsu": "No",
        "English-Medium Engineering": "Yes — GRADUATE GSC only",
        "Stipend Benchmark (JPY/month)": "¥200,000–300,000",
        "Risk Notes": "Over-courted student pool; UTokyo students may view 6-month internship as competing with research priorities.",
    },
    {
        "University": "Osaka University\n(大阪大学 / Handai)",
        "City": "Osaka",
        "Website": "https://www.osaka-u.ac.jp/en",
        "LinkedIn URL": "https://www.linkedin.com/school/osaka-university/",
        "Targeted Course / Collaborations": "Graduate School of Information Science and Technology; School of Engineering Science (Toyonaka, Information Science track) — English Master's tracks available but contracting",
        "No. of Recruiters\n(Competitive Intensity)": "High (regional)",
        "Feedback (Our Assessment)": "Top-3 national engineering brand and Kansai's strongest CS pipeline. English-medium engineering scope is contracting (Mechanical English program discontinued 2024) — verify currency before targeting.",
        "Contact Person": "Not publicly listed",
        "Email": "Via School of Engineering career section (school-level only)",
        "Phone": "+81-6-6877-5111 (main)",
        "Possibility of Collaboration": "Medium",
        "How (Collaboration Mechanism)": "Through School of Engineering career section + Career Center; verify English Master's program currency first via international office.",
        "Past Collaboration Evidence / Link": "Indian IT majors recruiting via Kansai offices (TCS, Infosys, Wipro alumni presence); IBM Japan, Accenture, Intel Japan placements documented. URL: https://www.eng.osaka-u.ac.jp/en/",
        "Fit Score (1-10)": 8,
        "Recommended Cycle": "Year 4+ (deprioritize)",
        "Best First-Contact Channel": "School of Engineering career section",
        "Language Preference": "JP primary",
        "Off-cycle from Shinsotsu": "No",
        "English-Medium Engineering": "Partial — contracting",
        "Stipend Benchmark (JPY/month)": "¥150,000–250,000",
        "Risk Notes": "English-medium engineering scope may continue to contract; if Mechanical was cut, others may follow.",
    },
    {
        "University": "Keio University\n(慶應義塾大学)",
        "City": "Tokyo / Yokohama",
        "Website": "https://www.keio.ac.jp/en/",
        "LinkedIn URL": "https://www.linkedin.com/school/keio-university/",
        "Targeted Course / Collaborations": "SFC (Faculty of Environment & Information Studies) — GIGA Program English-medium bachelor's; Faculty of Science & Technology Information & Computer Science dept (mostly Japanese-medium)",
        "No. of Recruiters\n(Competitive Intensity)": "Very High",
        "Feedback (Our Assessment)": "SFC's project/entrepreneurship culture is the right entry point for long-term internships. Faculty of S&T is more traditional shinsotsu-oriented. GIGA Program covers policy/environment/information — useful but not pure SDE.",
        "Contact Person": "Not publicly listed",
        "Email": "Not publicly listed for employer outreach",
        "Phone": "+81-3-5427-1517 (SFC main)",
        "Possibility of Collaboration": "Low–Medium for pure SDE",
        "How (Collaboration Mechanism)": "Engage via SFC faculty + project-based collaborations; not central career office. GIGA program coordinator for English-medium track.",
        "Past Collaboration Evidence / Link": "Strong alumni network at Mercari, SmartNews, Google Japan, McKinsey, BCG. URL: https://www.keio.ac.jp/en/student-life/career-support/",
        "Fit Score (1-10)": 7,
        "Recommended Cycle": "Year 3+ (only if scale needed)",
        "Best First-Contact Channel": "SFC-specific faculty / project route",
        "Language Preference": "JP primary",
        "Off-cycle from Shinsotsu": "SFC partial",
        "English-Medium Engineering": "Partial (GIGA non-engineering)",
        "Stipend Benchmark (JPY/month)": "¥200,000–300,000",
        "Risk Notes": "Faculty of S&T is traditional shinsotsu-oriented; SFC isn't pure SDE. May not yield SDE intern volume.",
    },
    {
        "University": "Nagoya Institute of Technology\n(名古屋工業大学 / NIT-Nagoya)",
        "City": "Nagoya",
        "Website": "https://www.nitech.ac.jp/eng/",
        "LinkedIn URL": "Limited public LinkedIn presence — [verify before use]",
        "Targeted Course / Collaborations": "Faculty of Engineering — Computer Science Dept; IGPGE (International Graduate Program for Global Engineers) English-medium Master's/PhD; joint doctoral program in Informatics with University of Wollongong",
        "No. of Recruiters\n(Competitive Intensity)": "Medium",
        "Feedback (Our Assessment)": "Smaller, technical-only institution with high signal-to-noise for SDE-leaning candidates. IGPGE cohort is globally minded. Less competitive recruiting market = easier entry for Celebal. Recent India-focused engagement events on record.",
        "Contact Person": "Not publicly listed (use IGPGE coordinator office)",
        "Email": "Through campus contact form; not centrally published",
        "Phone": "+81-52-735-5111 (main)",
        "Possibility of Collaboration": "Medium–High (low-volume)",
        "How (Collaboration Mechanism)": "Engage IGPGE coordinator + individual lab PIs; explore joint programs / cotutelle leverage. Smaller career office means relationship-led approach.",
        "Past Collaboration Evidence / Link": "Joint doctoral program with University of Wollongong (Australia) — proves international partnership capability. India-focused engagement events publicly noted.",
        "Fit Score (1-10)": 6,
        "Recommended Cycle": "Year 2 (Nagoya secondary feeder)",
        "Best First-Contact Channel": "IGPGE program coordinator office",
        "Language Preference": "JP primary; EN via IGPGE",
        "Off-cycle from Shinsotsu": "Partial (joint programs)",
        "English-Medium Engineering": "Partial (IGPGE graduate)",
        "Stipend Benchmark (JPY/month)": "¥150,000–220,000",
        "Risk Notes": "Smaller talent pool; lower brand recognition; smaller career office requires relationship-led outreach.",
    },
    {
        "University": "Osaka Metropolitan University\n(大阪公立大学 / OMU)",
        "City": "Osaka",
        "Website": "https://www.omu.ac.jp/en/",
        "LinkedIn URL": "Limited post-merger presence — [verify before use]",
        "Targeted Course / Collaborations": "Graduate School of Engineering, Graduate School of Informatics (both inherited from predecessor OCU/OPU); undergraduate engineering",
        "No. of Recruiters\n(Competitive Intensity)": "Medium (Kansai)",
        "Feedback (Our Assessment)": "Newest institution in the shortlist (post-2022 merger). Lowest English-program footprint. Useful only as a volume play; lowest English-readiness signal. Career-services brand recognition outside Japan is limited.",
        "Contact Person": "Not publicly listed",
        "Email": "Not publicly listed",
        "Phone": "Not publicly indexed in English",
        "Possibility of Collaboration": "Low",
        "How (Collaboration Mechanism)": "Generic career office routes; very limited English-language employer engagement infrastructure published.",
        "Past Collaboration Evidence / Link": "Kansai industry placements from predecessor universities (Panasonic-affiliated joint ventures, Intel Japan); no centralized published data. URL: https://www.omu.ac.jp/en/",
        "Fit Score (1-10)": 5,
        "Recommended Cycle": "Park (Year 4+ only)",
        "Best First-Contact Channel": "Generic career office (no published name)",
        "Language Preference": "JP primary",
        "Off-cycle from Shinsotsu": "No",
        "English-Medium Engineering": "Mostly Japanese",
        "Stipend Benchmark (JPY/month)": "¥150,000–250,000",
        "Risk Notes": "Post-merger administrative integration ongoing; career office contact churn risk; lower English-comfort baseline requires careful screening.",
    },
]


def main():
    wb = Workbook()

    # ---------------- Sheet 1: University Matrix ----------------
    ws = wb.active
    ws.title = "University Matrix"

    arial = Font(name="Arial", size=10)
    arial_bold = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    arial_bold_dark = Font(name="Arial", size=10, bold=True, color="000000")
    arial_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")

    header_fill = PatternFill("solid", start_color="1F4E78")
    title_fill = PatternFill("solid", start_color="2E75B6")
    high_fill = PatternFill("solid", start_color="C6EFCE")
    medium_fill = PatternFill("solid", start_color="FFEB9C")
    low_fill = PatternFill("solid", start_color="FFC7CE")
    anchor_fill = PatternFill("solid", start_color="D9E1F2")

    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="left")
    wrap_center_center = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value="Japan University Target Matrix — Celebal SDE / Developer Intern Program (2026Q2)")
    title_cell.font = arial_title
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title / context row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    sub = ws.cell(row=2, column=1, value="Sorted by recommended action priority. Strategic anchor: Nagoya University. See 'Notes & Legend' sheet for definitions.")
    sub.font = Font(name="Arial", size=10, italic=True, color="404040")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Header row
    header_row = 3
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = arial_bold
        cell.fill = header_fill
        cell.alignment = wrap_center_center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 32

    # Data rows
    for row_idx, row_data in enumerate(ROWS, start=header_row + 1):
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = arial
            cell.alignment = wrap_center
            cell.border = border

            # Color-code Possibility column
            if header == "Possibility of Collaboration":
                v = str(value).lower()
                if "high" in v and "low" not in v and "medium" not in v:
                    cell.fill = high_fill
                elif "medium" in v or "low–medium" in v or "low-medium" in v or "medium–high" in v or "medium-high" in v:
                    cell.fill = medium_fill
                elif "low" in v:
                    cell.fill = low_fill
                cell.alignment = wrap_center_center

            # Color-code Fit Score
            if header == "Fit Score (1-10)" and isinstance(value, int):
                if value >= 9:
                    cell.fill = high_fill
                elif value >= 7:
                    cell.fill = medium_fill
                else:
                    cell.fill = low_fill
                cell.alignment = wrap_center_center

            # Highlight anchor row (Nagoya, row 4)
            if row_idx == header_row + 1 and header == "University":
                cell.fill = anchor_fill
                cell.font = Font(name="Arial", size=10, bold=True)

            # Recommended Cycle visual
            if header == "Recommended Cycle":
                if "Year 1" in str(value):
                    cell.fill = high_fill
                    cell.font = Font(name="Arial", size=10, bold=True)
                elif "Year 2" in str(value):
                    cell.fill = medium_fill
                elif "Year 3" in str(value):
                    cell.fill = medium_fill
                else:
                    cell.fill = low_fill
                cell.alignment = wrap_center_center

        ws.row_dimensions[row_idx].height = 130

    # Freeze panes — keep header + first column visible
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # Auto filter on header
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(ROWS)}"

    # ---------------- Sheet 2: Notes & Legend ----------------
    legend = wb.create_sheet("Notes & Legend")

    legend.merge_cells("A1:C1")
    title = legend.cell(row=1, column=1, value="Notes & Legend")
    title.font = arial_title
    title.fill = title_fill
    title.alignment = Alignment(horizontal="center", vertical="center")
    legend.row_dimensions[1].height = 28

    notes = [
        ("Field", "Definition", "Notes"),
        ("Fit Score", "Composite 1–10 ranking", "30% historical yield (where available) + 25% course relevance & English-medium availability + 20% ranking/accreditation + 15% geographic & logistics + 10% diversity / international support"),
        ("Possibility of Collaboration", "High / Medium / Low / Low–Medium / Medium–High", "Reflects realistic 12-month outreach success probability given Japanese career-office norms, gaishi-kei competition, and Celebal brand recognition in Japan."),
        ("No. of Recruiters (Competitive Intensity)", "Very High / High / Medium / Low", "Estimated competitive intensity = how many companies actively recruit at the university for similar roles. Higher value = harder for Celebal to stand out."),
        ("Recommended Cycle", "Year 1 / Year 2 / Year 3+ / Year 4+ (deprioritize) / Park", "Year 1 = anchor for partnership MoU. Year 2+ = expand only after Year 1 produces results. Park = no action unless specific business trigger."),
        ("Off-cycle from Shinsotsu", "Yes / Partial / No", "Japan's April mass-hire (shinsotsu) makes off-cycle 6-month internships competitive with student job-hunt timing. Off-cycle availability is a significant de-risking factor."),
        ("English-Medium Engineering", "Yes (Undergrad) / Yes (Graduate) / Partial / Mostly Japanese", "G30 Undergrad is the strongest signal — Nagoya is uniquely positioned here."),
        ("Stipend Benchmark", "Range in JPY/month for engineering long-term interns at gaishi-kei firms in that region", "Floor for competitive offers; below this = candidates choose Google Japan / Indeed / Mercari instead."),
        ("Past Collaboration Evidence", "Public evidence of foreign-firm engagement", "We do NOT have access to confidential MoU registries. Evidence cited is from public alumni placements, news, university press releases."),
        ("Contact Person", "Named POC where publicly listed", "Most Japanese university career offices do not publish named POCs publicly. 'Not publicly listed' means we will not invent one."),
        ("LinkedIn URL", "Official university LinkedIn page", "Some Japanese universities have limited LinkedIn presence; URLs marked '[verify]' need confirmation before being used in stakeholder communications."),
    ]

    for r, row in enumerate(notes, start=3):
        for c, val in enumerate(row, start=1):
            cell = legend.cell(row=r, column=c, value=val)
            cell.font = arial_bold_dark if r == 3 else arial
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.border = border
            if r == 3:
                cell.fill = header_fill
                cell.font = arial_bold
        legend.row_dimensions[r].height = 50 if r > 3 else 28

    legend.column_dimensions["A"].width = 32
    legend.column_dimensions["B"].width = 32
    legend.column_dimensions["C"].width = 75

    # Color legend section
    color_section_start = 3 + len(notes) + 2
    legend.cell(row=color_section_start, column=1, value="Color Legend").font = arial_bold_dark
    color_legend = [
        ("High / Year 1 / Score ≥9", "C6EFCE"),
        ("Medium / Year 2-3 / Score 7-8", "FFEB9C"),
        ("Low / Park / Score ≤6", "FFC7CE"),
        ("Anchor University (Nagoya)", "D9E1F2"),
    ]
    for i, (label, color) in enumerate(color_legend, start=1):
        label_cell = legend.cell(row=color_section_start + i, column=1, value=label)
        label_cell.font = arial
        label_cell.border = border
        swatch = legend.cell(row=color_section_start + i, column=2, value="")
        swatch.fill = PatternFill("solid", start_color=color)
        swatch.border = border

    # Provenance
    prov_start = color_section_start + len(color_legend) + 3
    legend.cell(row=prov_start, column=1, value="Provenance").font = arial_bold_dark
    provenance = [
        ("Produced", "2026-05-26"),
        ("Produced by", "research-agent (general-purpose surrogate)"),
        ("Sources", "Public only — official university websites, MEXT publications, THE/QS rankings, Japanese press"),
        ("Target role", "Software Engineering / Developer interns (Python / Java / cloud / data)"),
        ("Companion docs", "japan-intern-shortlist-2026Q2.md; japan-intern-outreach-pack-2026Q2.md; japan-city-and-partnership-strategy-2026Q2.md"),
        ("Next review due", "2026-08-26 (3-month cadence)"),
    ]
    for i, (k, v) in enumerate(provenance, start=1):
        legend.cell(row=prov_start + i, column=1, value=k).font = arial_bold_dark
        legend.cell(row=prov_start + i, column=2, value=v).font = arial
        legend.cell(row=prov_start + i, column=1).border = border
        legend.cell(row=prov_start + i, column=2).border = border
        legend.merge_cells(start_row=prov_start + i, start_column=2, end_row=prov_start + i, end_column=3)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote: {OUTPUT}")


if __name__ == "__main__":
    main()
